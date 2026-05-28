# src/loaders/xlsx_loader.py

"""XLSX loader — one page per sheet."""

from __future__ import annotations

from pathlib import Path
from typing import List

from src.loaders.base import BaseLoader, LoadedDocument, LoadedPage, LoaderError


def _format_row(values: List) -> str:
    """Render one row of cell values as a ` | `-separated line."""
    return " | ".join("" if v is None else str(v).strip() for v in values)


class XLSXLoader(BaseLoader):
    name = "xlsx"
    extensions = (".xlsx", ".xlsm")
    mime_types = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    def load(self, path: Path) -> LoadedDocument:
        path = self._resolve(path)

        # Lazy import — openpyxl pulls in zipfile/xml machinery; only
        # cost it when a workbook is actually being loaded.
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency missing
            raise LoaderError(f"openpyxl is required to load {path}") from exc

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise LoaderError(f"Could not open xlsx {path}: {exc}") from exc

        pages: List[LoadedPage] = []
        total_rows = 0
        for page_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            lines: List[str] = []
            sheet_rows = 0
            for row in ws.iter_rows(values_only=True):
                # Skip totally-empty rows; they bloat text without value.
                if not any(v not in (None, "") for v in row):
                    continue
                lines.append(_format_row(list(row)))
                sheet_rows += 1
            total_rows += sheet_rows

            meta = self._base_page_metadata(path, page_number=page_idx)
            meta["sheet_name"] = sheet_name
            meta["row_count"] = sheet_rows
            pages.append(LoadedPage(text="\n".join(lines), page_number=page_idx, metadata=meta))

        wb.close()

        return LoadedDocument(
            source_path=str(path),
            document_name=path.name,
            mime_type=self._mime_for(path) or self.mime_types[0],
            pages=pages,
            metadata={"sheet_count": len(pages), "total_rows": total_rows},
        )
